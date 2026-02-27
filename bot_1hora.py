"""
=============================================================
  BOT 1 HORA - CEDEARs en reMarkets (Primary API)
  Timeframe: Velas de 1 hora | Historial: 3 meses
  Stop Loss: 2% | Take Profit: 5%
  Señal: 4 de 5 condiciones + EMA200 + Filtro SPY
  Trailing Stop dinamico
  Cierre por tiempo: 10 dias
  Horario: 11:00 a 16:00 hs
  Autor: Generado con Claude - Anthropic
=============================================================

INSTALACION:
    pip install requests pandas ta python-dotenv schedule yfinance openpyxl

CONFIGURACION (.env):
    REMARKETS_USER=tu_usuario
    REMARKETS_PASSWORD=tu_password
    REMARKETS_ACCOUNT=tu_numero_de_cuenta

EJECUCION:
    python bot_1hora.py
=============================================================
"""

import os
import time
import logging
import requests
import pandas as pd
import ta
import yfinance as yf
from datetime import datetime, timedelta
from dotenv import load_dotenv
import schedule
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

load_dotenv()

# ─────────────────────────────────────────────
#  CONFIGURACION
# ─────────────────────────────────────────────

REMARKETS_USER     = os.getenv("REMARKETS_USER")
REMARKETS_PASSWORD = os.getenv("REMARKETS_PASSWORD")
REMARKETS_ACCOUNT  = os.getenv("REMARKETS_ACCOUNT")
BASE_URL           = "https://api.remarkets.primary.com.ar"

CAPITAL_TOTAL      = 1_000_000
CAPITAL_POR_OP     = 100_000

STOP_LOSS_PCT      = 0.02
TAKE_PROFIT_PCT    = 0.05
COMISION_IDA_VUELTA = 0.0044   # 0.17% + 0.05% x2 (con intraday bonificado)
MAX_DIAS_ABIERTO   = 10

HORARIO_INICIO     = 11
HORARIO_FIN        = 17
INTERVALO_MINUTOS  = 5

PAPER_TRADING      = True      # ⚠️ Cambia a False para dinero real

EXCEL_PATH         = "registro_operaciones.xlsx"
BOT_NOMBRE         = "1HORA"

CEDEARS = [
    "AAPL","AMD","AMZN","BABA","DIS","GOOGL","INTC","KO","JPM","MA",
    "MELI","META","MSFT","NFLX","NIO","NKE","NOK","NVDA","PBR","PEP",
    "PYPL","TSLA","T","TTWO","V","WMT","XOM"
]

TICKERS_BYMA = {c: f"MERV - XMEV - {c} - CI" for c in CEDEARS}

# ─────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler("bot_1hora.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
#  ESTADO EN MEMORIA
# ─────────────────────────────────────────────

posiciones         = {}
capital_disponible = CAPITAL_TOTAL
operaciones_total  = 0
ganancias_total    = 0.0
auth_token         = None
op_id_contador     = 1

# ─────────────────────────────────────────────
#  AUTENTICACION
# ─────────────────────────────────────────────

def autenticar() -> bool:
    global auth_token
    try:
        resp = requests.post(
            f"{BASE_URL}/auth/getToken",
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "X-Username":   REMARKETS_USER,
                "X-Password":   REMARKETS_PASSWORD
            },
            timeout=10
        )
        resp.raise_for_status()
        auth_token = resp.headers.get("X-Auth-Token")
        if not auth_token:
            log.error("No se recibio X-Auth-Token")
            return False
        log.info("Autenticacion exitosa en reMarkets")
        return True
    except Exception as e:
        log.error(f"Error al autenticar: {e}")
        return False

def headers_auth():
    return {"Content-Type": "application/json; charset=utf-8", "X-Auth-Token": auth_token}

# ─────────────────────────────────────────────
#  DATOS DE MERCADO
# ─────────────────────────────────────────────

def obtener_historial(ticker: str) -> pd.DataFrame:
    """Velas de 1 hora, 3 meses de historial desde Yahoo Finance."""
    try:
        df = yf.download(ticker, period="3mo", interval="1h", progress=False, auto_adjust=True)
        if df.empty:
            return pd.DataFrame()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.columns = [c.lower() for c in df.columns]
        return df[["open","high","low","close","volume"]].dropna()
    except Exception as e:
        log.error(f"Error historial {ticker}: {e}")
        return pd.DataFrame()

def obtener_precio_actual(cedear: str) -> float:
    """Precio en tiempo real desde reMarkets, fallback Yahoo."""
    try:
        resp = requests.get(
            f"{BASE_URL}/rest/marketdata/get",
            params={"marketId": "ROFX", "symbol": TICKERS_BYMA[cedear], "entries": "LA,OF", "depth": 1},
            headers=headers_auth(),
            timeout=10
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("status") == "OK":
            md = data.get("marketData", {})
            la = md.get("LA")
            if la and la.get("price"):
                return float(la["price"])
            of = md.get("OF")
            if of and len(of) > 0:
                return float(of[0]["price"])
    except Exception:
        pass
    try:
        df = yf.download(cedear, period="1d", interval="1m", progress=False, auto_adjust=True)
        return float(df["Close"].iloc[-1]) if not df.empty else 0.0
    except:
        return 0.0

def verificar_spy_alcista() -> bool:
    """Filtro mercado general: SPY debe estar por encima de su EMA 200."""
    try:
        df = yf.download("SPY", period="1y", interval="1d", progress=False, auto_adjust=True)
        if df.empty or len(df) < 200:
            return True
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.columns = [c.lower() for c in df.columns]
        close = df["close"].astype(float)
        ema200 = ta.trend.EMAIndicator(close, window=200).ema_indicator()
        return float(close.iloc[-1]) > float(ema200.iloc[-1])
    except:
        return True

# ─────────────────────────────────────────────
#  CALCULO DE INDICADORES
# ─────────────────────────────────────────────

def calcular_indicadores(df: pd.DataFrame) -> dict:
    if df.empty or len(df) < 50:
        return {}
    try:
        close = df["close"].astype(float)
        vol   = df["volume"].astype(float)

        rsi   = ta.momentum.RSIIndicator(close, window=14).rsi()
        ema9  = ta.trend.EMAIndicator(close, window=9).ema_indicator()
        ema21 = ta.trend.EMAIndicator(close, window=21).ema_indicator()
        ema200 = ta.trend.EMAIndicator(close, window=200).ema_indicator() if len(df) >= 200 else None

        macd_obj    = ta.trend.MACD(close, window_fast=12, window_slow=26, window_sign=9)
        macd_line   = macd_obj.macd()
        macd_signal = macd_obj.macd_signal()

        bb       = ta.volatility.BollingerBands(close, window=20, window_dev=2)
        bb_lower = bb.bollinger_lband()
        bb_upper = bb.bollinger_hband()

        vol_medio = vol.rolling(20).mean()

        df2 = pd.DataFrame({
            "close": close, "rsi": rsi,
            "ema9": ema9, "ema21": ema21,
            "macd": macd_line, "macd_sig": macd_signal,
            "bb_lower": bb_lower, "bb_upper": bb_upper,
            "volume": vol, "vol_medio": vol_medio,
        }).dropna()

        if ema200 is not None:
            df2["ema200"] = ema200.reindex(df2.index)

        if len(df2) < 2:
            return {}

        u = df2.iloc[-1]
        a = df2.iloc[-2]

        return {
            "precio":          float(u["close"]),
            "rsi":             float(u["rsi"]),
            "ema9":            float(u["ema9"]),
            "ema21":           float(u["ema21"]),
            "ema9_ant":        float(a["ema9"]),
            "ema21_ant":       float(a["ema21"]),
            "ema200":          float(u["ema200"]) if "ema200" in u else None,
            "macd":            float(u["macd"]),
            "macd_signal":     float(u["macd_sig"]),
            "macd_ant":        float(a["macd"]),
            "macd_signal_ant": float(a["macd_sig"]),
            "bb_lower":        float(u["bb_lower"]),
            "bb_upper":        float(u["bb_upper"]),
            "volumen":         float(u["volume"]),
            "vol_medio":       float(u["vol_medio"]),
        }
    except Exception as e:
        log.error(f"Error indicadores: {e}")
        return {}

# ─────────────────────────────────────────────
#  LOGICA DE SEÑALES
# ─────────────────────────────────────────────

def evaluar_señal_compra(ind: dict) -> bool:
    """4 de 5 condiciones + filtro EMA 200."""
    if not ind:
        return False
    if ind["ema200"] and ind["precio"] < ind["ema200"]:
        return False

    cruce_ema_alcista  = ind["ema9_ant"] <= ind["ema21_ant"] and ind["ema9"] > ind["ema21"]
    cruce_macd_alcista = ind["macd_ant"] <= ind["macd_signal_ant"] and ind["macd"] > ind["macd_signal"]
    cerca_bb_inferior  = ind["precio"] <= ind["bb_lower"] * 1.01
    volumen_alto       = ind["volumen"] > ind["vol_medio"] * 1.2

    condiciones = sum([
        ind["rsi"] < 35,
        cruce_ema_alcista,
        cruce_macd_alcista,
        cerca_bb_inferior,
        volumen_alto
    ])

    return condiciones >= 4

def evaluar_señal_venta(ind: dict) -> bool:
    """3 de 4 condiciones de salida técnica."""
    if not ind:
        return False

    cruce_ema_bajista  = ind["ema9_ant"] >= ind["ema21_ant"] and ind["ema9"] < ind["ema21"]
    cruce_macd_bajista = ind["macd_ant"] >= ind["macd_signal_ant"] and ind["macd"] < ind["macd_signal"]
    cerca_bb_superior  = ind["precio"] >= ind["bb_upper"] * 0.99

    condiciones = sum([
        ind["rsi"] > 65,
        cruce_ema_bajista,
        cruce_macd_bajista,
        cerca_bb_superior
    ])

    return condiciones >= 3

# ─────────────────────────────────────────────
#  TRAILING STOP
# ─────────────────────────────────────────────

def actualizar_trailing_stop(cedear: str, precio_actual: float):
    pos = posiciones[cedear]
    if precio_actual > pos["trailing_max"]:
        pos["trailing_max"] = precio_actual
        nuevo_sl = round(precio_actual * (1 - STOP_LOSS_PCT), 2)
        if nuevo_sl > pos["stop_loss"]:
            pos["stop_loss"] = nuevo_sl
            log.info(f"  Trailing SL actualizado {cedear}: ${nuevo_sl:.2f}")

# ─────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────

def registrar_en_excel(cedear, fecha_entrada, hora_entrada, precio_entrada,
                        cantidad, capital, fecha_salida, hora_salida,
                        precio_salida, motivo):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Operaciones"]
        fila = 4
        while ws.cell(row=fila, column=2).value is not None:
            fila += 1

        ws.cell(row=fila, column=2,  value=cedear)
        ws.cell(row=fila, column=3,  value=BOT_NOMBRE)
        ws.cell(row=fila, column=4,  value=fecha_entrada)
        ws.cell(row=fila, column=5,  value=hora_entrada)
        ws.cell(row=fila, column=6,  value=precio_entrada)
        ws.cell(row=fila, column=7,  value=cantidad)
        ws.cell(row=fila, column=8,  value=capital)
        ws.cell(row=fila, column=9,  value=fecha_salida)
        ws.cell(row=fila, column=10, value=hora_salida)
        ws.cell(row=fila, column=11, value=precio_salida)
        ws.cell(row=fila, column=12, value=motivo)

        resultado = (precio_salida - precio_entrada) * cantidad
        color = "E2EFDA" if resultado >= 0 else "FFDCDC"
        fill = PatternFill("solid", start_color=color)
        for col in range(1, 19):
            ws.cell(row=fila, column=col).fill = fill

        wb.save(EXCEL_PATH)
        log.info(f"  Operación registrada en Excel fila {fila}")
    except Exception as e:
        log.warning(f"  No se pudo escribir en Excel: {e}")

# ─────────────────────────────────────────────
#  GESTION DE ORDENES
# ─────────────────────────────────────────────

def ejecutar_orden(cedear: str, tipo: str, cantidad: int, precio: float) -> bool:
    if PAPER_TRADING:
        log.info(f"  [PAPER] {tipo} {cantidad}x {cedear} @ ${precio:.2f}")
        return True
    try:
        resp = requests.get(
            f"{BASE_URL}/rest/order/newSingleOrder",
            params={
                "marketId":       "ROFX",
                "symbol":         TICKERS_BYMA[cedear],
                "side":           tipo,
                "orderQty":       cantidad,
                "ordType":        "Market",
                "timeInForce":    "Day",
                "account":        REMARKETS_ACCOUNT,
                "cancelPrevious": "False",
                "iceberg":        "False"
            },
            headers=headers_auth(),
            timeout=10
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("status") == "OK":
            log.info(f"  Orden OK: {tipo} {cantidad}x {cedear} @ ${precio:.2f}")
            return True
        log.error(f"  Orden rechazada: {data}")
        return False
    except Exception as e:
        log.error(f"  Error orden {tipo} {cedear}: {e}")
        if "401" in str(e) or "403" in str(e):
            autenticar()
        return False

def cerrar_posicion(cedear: str, precio: float, motivo: str):
    global capital_disponible, operaciones_total, ganancias_total
    pos = posiciones[cedear]
    if ejecutar_orden(cedear, "Sell", pos["cantidad"], precio):
        ganancia = (precio - pos["precio_entrada"]) * pos["cantidad"]
        comision = pos["capital"] * COMISION_IDA_VUELTA
        neto     = ganancia - comision
        capital_disponible += precio * pos["cantidad"]
        ganancias_total    += neto
        operaciones_total  += 1

        log.info(
            f"  CIERRE {cedear} | Motivo: {motivo} | "
            f"Entrada: ${pos['precio_entrada']:.2f} → Salida: ${precio:.2f} | "
            f"Resultado: ${ganancia:+.2f} | Neto: ${neto:+.2f} ARS"
        )

        registrar_en_excel(
            cedear=cedear,
            fecha_entrada=pos["fecha_entrada"].strftime("%d/%m/%Y"),
            hora_entrada=pos["fecha_entrada"].strftime("%H:%M"),
            precio_entrada=pos["precio_entrada"],
            cantidad=pos["cantidad"],
            capital=pos["capital"],
            fecha_salida=datetime.now().strftime("%d/%m/%Y"),
            hora_salida=datetime.now().strftime("%H:%M"),
            precio_salida=precio,
            motivo=motivo
        )
        del posiciones[cedear]

# ─────────────────────────────────────────────
#  LOGICA PRINCIPAL POR CEDEAR
# ─────────────────────────────────────────────

def procesar_cedear(cedear: str, spy_alcista: bool):
    global capital_disponible

    df = obtener_historial(cedear)
    if df.empty:
        return

    ind = calcular_indicadores(df)
    if not ind:
        return

    precio = obtener_precio_actual(cedear)
    if precio == 0.0:
        precio = ind["precio"]

    posicion = posiciones.get(cedear)

    if posicion:
        actualizar_trailing_stop(cedear, precio)
        dias_abierto = (datetime.now() - posicion["fecha_entrada"]).days

        if precio <= posicion["stop_loss"]:
            cerrar_posicion(cedear, precio, "STOP LOSS")
            return

        if precio >= posicion["take_profit"]:
            cerrar_posicion(cedear, precio, "TAKE PROFIT")
            return

        if dias_abierto >= MAX_DIAS_ABIERTO:
            cerrar_posicion(cedear, precio, f"TIEMPO ({dias_abierto} dias)")
            return

        if evaluar_señal_venta(ind):
            cerrar_posicion(cedear, precio, "SEÑAL VENTA")
            return

        log.info(
            f"  {cedear} | ${precio:.2f} | RSI:{ind['rsi']:.0f} | "
            f"SL:${posicion['stop_loss']:.2f} | TP:${posicion['take_profit']:.2f} | "
            f"Dias:{dias_abierto}"
        )
        return

    if not spy_alcista:
        return

    if capital_disponible < CAPITAL_POR_OP:
        return

    if evaluar_señal_compra(ind):
        cantidad = int(CAPITAL_POR_OP / precio)
        if cantidad <= 0:
            return
        costo = precio * cantidad

        if ejecutar_orden(cedear, "Buy", cantidad, precio):
            capital_disponible -= costo
            stop_loss   = round(precio * (1 - STOP_LOSS_PCT), 2)
            take_profit = round(precio * (1 + TAKE_PROFIT_PCT), 2)
            posiciones[cedear] = {
                "precio_entrada": precio,
                "cantidad":       cantidad,
                "capital":        costo,
                "stop_loss":      stop_loss,
                "take_profit":    take_profit,
                "trailing_max":   precio,
                "fecha_entrada":  datetime.now()
            }
            log.info(
                f"  COMPRA {cedear} | {cantidad} acc x ${precio:.2f} = ${costo:.2f} ARS | "
                f"SL:${stop_loss:.2f} | TP:${take_profit:.2f}"
            )

# ─────────────────────────────────────────────
#  RESUMEN
# ─────────────────────────────────────────────

def mostrar_resumen():
    log.info("=" * 65)
    log.info(f"[BOT 1HORA] PORTFOLIO - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    log.info(f"  Capital disponible : ${capital_disponible:,.2f} ARS")
    log.info(f"  Posiciones abiertas: {len(posiciones)}")
    log.info(f"  Operaciones cerradas: {operaciones_total}")
    log.info(f"  Resultado acumulado: ${ganancias_total:+,.2f} ARS")
    for cedear, pos in posiciones.items():
        precio_actual = obtener_precio_actual(cedear)
        pnl = (precio_actual - pos["precio_entrada"]) * pos["cantidad"]
        pnl_pct = ((precio_actual / pos["precio_entrada"]) - 1) * 100
        log.info(f"  {cedear}: ${pos['precio_entrada']:.2f}→${precio_actual:.2f} | PnL: ${pnl:+.2f} ({pnl_pct:+.1f}%)")
    log.info("=" * 65)

# ─────────────────────────────────────────────
#  CICLO PRINCIPAL
# ─────────────────────────────────────────────

def ciclo_trading():
    ahora = datetime.now()
    if ahora.weekday() >= 5:
        return
    if not (HORARIO_INICIO <= ahora.hour < HORARIO_FIN):
        return

    log.info(f"--- Ciclo {ahora.strftime('%H:%M')} ---")

    spy_alcista = verificar_spy_alcista()
    if not spy_alcista:
        log.warning("  SPY por debajo de EMA200 - no se abren nuevas posiciones")

    for cedear in CEDEARS:
        try:
            procesar_cedear(cedear, spy_alcista)
            time.sleep(1)
        except Exception as e:
            log.error(f"  Error procesando {cedear}: {e}")

    mostrar_resumen()

def main():
    log.info("=" * 65)
    log.info("  BOT 1 HORA - CEDEARs reMarkets")
    log.info(f"  Capital: ${CAPITAL_TOTAL:,} ARS | Por op: ${CAPITAL_POR_OP:,} ARS")
    log.info(f"  Modo: {'PAPER TRADING' if PAPER_TRADING else '💰 REAL'}")
    log.info(f"  CEDEARs: {len(CEDEARS)} instrumentos")
    log.info(f"  SL: {STOP_LOSS_PCT*100:.0f}% | TP: {TAKE_PROFIT_PCT*100:.0f}% | Trailing: SI")
    log.info(f"  Señal: 4 de 5 condiciones + EMA200 + Filtro SPY")
    log.info(f"  Cierre por tiempo: {MAX_DIAS_ABIERTO} dias")
    log.info(f"  Horario: {HORARIO_INICIO}:00 a {HORARIO_FIN}:00 hs")
    log.info(f"  Comision estimada: {COMISION_IDA_VUELTA*100:.2f}% ida y vuelta (intraday bonificado)")
    log.info("=" * 65)

    if not autenticar():
        log.error("No se pudo autenticar. Verificá el .env")
        return

    ciclo_trading()
    schedule.every(INTERVALO_MINUTOS).minutes.do(ciclo_trading)

    while True:
        schedule.run_pending()
        time.sleep(30)

if __name__ == "__main__":
    main()
